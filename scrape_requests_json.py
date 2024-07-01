import requests
from fake_useragent import UserAgent
import random
import time
import sys
import os
from openpyxl import Workbook, load_workbook

# 主要是用來練習 requests 爬取 api 時的方法

# 獲取 json 資料
def get_fetch_json(url, session, headers):
    try:
        response = session.get(url, headers=headers)
        response.raise_for_status()
        return response.json()  # 返回渲染後的 HTML 內容
    except Exception as e:
        error_line = sys.exc_info()[-1].tb_lineno
        print(f"錯誤發生在第 {error_line} 行：\n{str(e)}")
        return None

# 解析 list json 資料，並組成 detail api url
def parse_list_json(datas):
    try:
        detail_api_urls = []
        
        for data in datas:
            if data['score'] is not None:
                api_url = "https://spa1.scrape.center/api/movie/%s/" %data['id']
                detail_api_urls.append(api_url)
            
        time.sleep(random.uniform(3, 5)) 
        return detail_api_urls
    except Exception as e:
        error_line = sys.exc_info()[-1].tb_lineno
        print(f"錯誤發生在第 {error_line} 行：\n{str(e)}")
        return None
    
# 解析 list json 資料，並組成 detail api url
def parse_detail_json(data):
    try:
        detail_data = []
        
        # 電影名稱
        detail_data.append(data['name'])
        # 上映日期
        detail_data.append(data['published_at'])
        # 電影評分
        detail_data.append(data['score'])
        # 電影時長
        detail_data.append(data['minute'] + " 分钟")
        # 電影簡介
        detail_data.append(data['drama'])
        
        time.sleep(random.uniform(3, 5)) 
        
        return detail_data
    except Exception as e:
        error_line = sys.exc_info()[-1].tb_lineno
        print(f"錯誤發生在第 {error_line} 行：\n{str(e)}")
        return None

# 將資料寫入 Excel 中 
def write_to_excel(data, file_name="scrape_requests_json.xlsx"):
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["電影名稱", "上映日期", "電影評分", "電影時長", "電影簡介"])
    
    duplicate_found = False   
    for row in sheet.iter_rows(values_only=True):
        if row[0] == data[0]:
            duplicate_found = True
            break
        
    if not duplicate_found:
        sheet.append(data)
        
    workbook.save(file_name)
        
    

if __name__ == "__main__":
    session_requests = requests.session()
    ua = UserAgent()
    headers = {
        'User-Agent': ua.random
    }
    
    offset = 0
    for page in range(1,999):
        list_url = f"https://spa1.scrape.center/api/movie/?limit=10&offset={offset*10}"
        print(f"正在爬取電影數據網站第 {page} 頁")
        list_json = get_fetch_json(list_url, session_requests, headers)
        offset += 1
        detail_urls = parse_list_json(list_json['results'])
        if detail_urls != []:
            for detail_url in detail_urls:
                print(f"正在爬取電影數據網站 {detail_url}")
                detail_json = get_fetch_json(detail_url, session_requests, headers)
                detail_data = parse_detail_json(detail_json)
                write_to_excel(detail_data)
        else:
            print(f"電影數據網站第 {page-1} 頁是最後一頁囉")
            break